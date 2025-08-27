import logging
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from db_manager import db_manager
from config import REPORTS_DIR, COMPANY_NAME, COMPANY_ADDRESS, COMPANY_PHONE

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ReportManager:
    def __init__(self):
        self.db = db_manager
        # Créer le dossier des rapports s'il n'existe pas
        if not os.path.exists(REPORTS_DIR):
            os.makedirs(REPORTS_DIR)
    
    def generate_excel_report(self, report_type, start_date, end_date, employee_id=None, department_id=None):
        """Générer un rapport Excel"""
        try:
            # Récupérer les données
            if report_type == 'daily':
                data = self._get_daily_data(start_date, end_date, employee_id, department_id)
                filename = f"rapport_quotidien_{start_date}_{end_date}.xlsx"
            elif report_type == 'monthly':
                data = self._get_monthly_data(start_date, end_date, employee_id, department_id)
                filename = f"rapport_mensuel_{start_date}_{end_date}.xlsx"
            else:
                logger.error(f"Type de rapport non supporté: {report_type}")
                return None
            
            file_path = os.path.join(REPORTS_DIR, filename)
            
            # Créer le classeur Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport de Présence"
            
            # En-tête du rapport
            ws.merge_cells('A1:H1')
            ws['A1'] = f"Rapport de Présence - {COMPANY_NAME}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:H2')
            ws['A2'] = f"Période: {start_date} au {end_date}"
            ws['A2'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A3:H3')
            ws['A3'] = f"Généré le: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # En-têtes du tableau
            headers = ['Matricule', 'Nom', 'Prénom', 'Département', 'Date', 'Heure Entrée', 'Heure Sortie', 'Heures Travail']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Remplir les données
            row = 6
            for record in data:
                ws.cell(row=row, column=1, value=record['employee_id'])
                ws.cell(row=row, column=2, value=record['last_name'])
                ws.cell(row=row, column=3, value=record['first_name'])
                ws.cell(row=row, column=4, value=record['department_name'])
                ws.cell(row=row, column=5, value=record['date'])
                ws.cell(row=row, column=6, value=record['time_in'])
                ws.cell(row=row, column=7, value=record['time_out'])
                ws.cell(row=row, column=8, value=record['work_hours'])
                row += 1
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder le fichier
            wb.save(file_path)
            logger.info(f"Rapport Excel généré: {file_path}")
            
            # Enregistrer dans la base de données
            self.db.add_report(report_type, start_date, end_date, file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"Erreur lors de la génération du rapport Excel: {e}")
            return None
    
    def generate_pdf_report(self, report_type, start_date, end_date, employee_id=None, department_id=None):
        """Générer un rapport PDF"""
        try:
            # Récupérer les données
            if report_type == 'daily':
                data = self._get_daily_data(start_date, end_date, employee_id, department_id)
                filename = f"rapport_quotidien_{start_date}_{end_date}.pdf"
            elif report_type == 'monthly':
                data = self._get_monthly_data(start_date, end_date, employee_id, department_id)
                filename = f"rapport_mensuel_{start_date}_{end_date}.pdf"
            else:
                logger.error(f"Type de rapport non supporté: {report_type}")
                return None
            
            file_path = os.path.join(REPORTS_DIR, filename)
            
            # Créer le document PDF
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Titre
            title = Paragraph(f"Rapport de Présence - {COMPANY_NAME}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Informations de l'entreprise
            company_info = Paragraph(f"{COMPANY_ADDRESS}<br/>Tél: {COMPANY_PHONE}", styles['Normal'])
            story.append(company_info)
            story.append(Spacer(1, 12))
            
            # Période du rapport
            period = Paragraph(f"Période: {start_date} au {end_date}", styles['Normal'])
            story.append(period)
            
            date_generated = Paragraph(f"Généré le: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            story.append(date_generated)
            story.append(Spacer(1, 24))
            
            # Tableau des données
            table_data = [['Matricule', 'Nom', 'Prénom', 'Département', 'Date', 'Entrée', 'Sortie', 'Heures']]
            
            for record in data:
                table_data.append([
                    record['employee_id'],
                    record['last_name'],
                    record['first_name'],
                    record['department_name'],
                    record['date'],
                    record['time_in'],
                    record['time_out'],
                    record['work_hours']
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 24))
            
            # Générer le PDF
            doc.build(story)
            logger.info(f"Rapport PDF généré: {file_path}")
            
            # Enregistrer dans la base de données
            self.db.add_report(report_type, start_date, end_date, file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"Erreur lors de la génération du rapport PDF: {e}")
            return None
    
    def _get_daily_data(self, start_date, end_date, employee_id=None, department_id=None):
        """Récupérer les données pour le rapport quotidien"""
        try:
            logs = self.db.get_attendance_logs(start_date, end_date, employee_id, department_id)
            
            # Organiser les données par employé et par date
            employee_data = {}
            for log in logs:
                emp_key = f"{log['employee_id']}_{log['date']}"
                if emp_key not in employee_data:
                    employee_data[emp_key] = {
                        'employee_id': log['employee_id'],
                        'first_name': log['first_name'],
                        'last_name': log['last_name'],
                        'department_name': log['department_name'],
                        'date': log['date'],
                        'time_in': None,
                        'time_out': None
                    }
                
                if log['type'] == 'IN':
                    employee_data[emp_key]['time_in'] = log['time']
                elif log['type'] == 'OUT':
                    employee_data[emp_key]['time_out'] = log['time']
            
            # Calculer les heures de travail
            result = []
            for emp_key, data in employee_data.items():
                if data['time_in'] and data['time_out']:
                    # Calculer les heures travaillées
                    in_time = datetime.strptime(f"{data['date']} {data['time_in']}", '%Y-%m-%d %H:%M:%S')
                    out_time = datetime.strptime(f"{data['date']} {data['time_out']}", '%Y-%m-%d %H:%M:%S')
                    work_hours = round((out_time - in_time).total_seconds() / 3600, 2)
                    
                    result.append({
                        'employee_id': data['employee_id'],
                        'first_name': data['first_name'],
                        'last_name': data['last_name'],
                        'department_name': data['department_name'],
                        'date': data['date'],
                        'time_in': data['time_in'],
                        'time_out': data['time_out'],
                        'work_hours': work_hours
                    })
            
            return result
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des données quotidiennes: {e}")
            return []
    
    def _get_monthly_data(self, start_date, end_date, employee_id=None, department_id=None):
        """Récupérer les données pour le rapport mensuel"""
        try:
            # Pour le rapport mensuel, on agrège les données par employé
            logs = self.db.get_attendance_logs(start_date, end_date, employee_id, department_id)
            
            employee_stats = {}
            for log in logs:
                emp_id = log['employee_id']
                if emp_id not in employee_stats:
                    employee_stats[emp_id] = {
                        'employee_id': log['employee_id'],
                        'first_name': log['first_name'],
                        'last_name': log['last_name'],
                        'department_name': log['department_name'],
                        'total_days': 0,
                        'total_hours': 0,
                        'overtime_hours': 0
                    }
                
                # Compter les jours de présence (au moins une entrée)
                if log['type'] == 'IN':
                    employee_stats[emp_id]['total_days'] += 1
                
                # Calculer les heures (nécessite une paire IN/OUT complète)
                # Cette logique serait plus complexe dans une implémentation réelle
            
            return list(employee_stats.values())
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des données mensuelles: {e}")
            return []

# Instance globale du gestionnaire de rapports
report_manager = ReportManager()
